from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from .models import Currency, CurrencyExchangeRate


@admin.register(Currency)
class CurrencyAdmin(admin.ModelAdmin):
    list_display = ['code', 'name']
    search_fields = ['code', 'name']


@admin.register(CurrencyExchangeRate)
class CurrencyExchangeRateAdmin(admin.ModelAdmin):
    list_display = ['base_currency', 'target_currency', 'rate', 'timestamp']
    list_filter = ['base_currency', 'target_currency', 'timestamp']
    search_fields = ['base_currency__code', 'target_currency__code']
    date_hierarchy = 'timestamp'

    actions = ['export_to_xlsx']

    def export_to_xlsx(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exchange Rates"

        headers = ['Base Currency', 'Target Currency', 'Rate', 'Timestamp']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, rate in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=rate.base_currency.code)
            ws.cell(row=row_num, column=2, value=rate.target_currency.code)
            ws.cell(row=row_num, column=3, value=float(rate.rate))
            ws.cell(row=row_num, column=4,
                    value=rate.timestamp.strftime('%Y-%m-%d %H:%M:%S'))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=exchange_rates.xlsx'
        wb.save(response)
        return response

    export_to_xlsx.short_description = "Export selected rates to XLSX"
